from __future__ import annotations

import hashlib
import json
import logging
import re
import time
import uuid
from dataclasses import dataclass
from pathlib import Path
from threading import Lock

import numpy as np


from .config import settings
from .contracts.tabular_contract import TabularSnapshot, build_snapshot, load_contract, validate_snapshot
from .embeddings import Embedder, HashEmbedder, NoEmbedder, SentenceTransformerEmbedder
from .index_maintenance import ensure_index_compatible
from .metadata import normalize_classification, normalize_retention, normalize_tags
from .ocr import extract_text_from_pdf
from .storage import (
    Chunk,
    IngestEvent,
    connect,
    delete_doc_contents,
    get_doc,
    init_db,
    insert_chunks,
    insert_embeddings,
    insert_ingest_event,
    list_ingest_events,
    upsert_doc,
)


def _vec_to_pgvector_literal(vec: np.ndarray) -> str:
    """Convert a 1D numpy vector to pgvector text format: "[1,2,3]"."""
    v = vec.astype(np.float32).reshape(-1)
    n = float(np.linalg.norm(v))
    if n > 0:
        v = v / n
    vals = v.tolist()
    return "[" + ",".join(str(float(x)) for x in vals) + "]"


logger = logging.getLogger(__name__)


# ---- tabular ingestion limits (CSV/TSV/XLSX)
#
# These are intentionally conservative:
# - uploads are already size-limited at the API layer
# - tabular files can expand when rendered as text
# - we want predictable chunk counts + latency
_TABULAR_MAX_ROWS = 2_000
_TABULAR_MAX_COLS = 60
_TABULAR_MAX_CELL_CHARS = 240
_TABULAR_MAX_TOTAL_CHARS = 1_000_000


def _slugify(text: str) -> str:
    t = re.sub(r"[^a-zA-Z0-9]+", "-", text.strip().lower()).strip("-")
    return t or "doc"


def stable_doc_id(title: str, source: str) -> str:
    """Short stable id derived from title + source."""
    h = hashlib.sha1(f"{title}\n{source}".encode("utf-8")).hexdigest()[:10]
    return f"{_slugify(title)[:32]}-{h}"


def chunk_text(text: str, chunk_size_chars: int, chunk_overlap_chars: int) -> list[str]:
    text = text.replace("\r\n", "\n")
    paras = [p.strip() for p in text.split("\n\n") if p.strip()]
    chunks: list[str] = []
    buf = ""

    def flush() -> None:
        nonlocal buf
        if buf.strip():
            chunks.append(buf.strip())
        buf = ""

    for p in paras:
        if len(buf) + len(p) + 2 <= chunk_size_chars:
            buf = f"{buf}\n\n{p}".strip()
        else:
            flush()
            # If a single paragraph is huge, hard split it
            while len(p) > chunk_size_chars:
                chunks.append(p[:chunk_size_chars].strip())
                p = p[chunk_size_chars - chunk_overlap_chars :]
            buf = p.strip()

    flush()

    # Add overlap by simple char overlap between adjacent chunks
    if chunk_overlap_chars > 0 and len(chunks) > 1:
        overlapped: list[str] = []
        prev_tail = ""
        for c in chunks:
            if prev_tail:
                overlapped.append((prev_tail + "\n" + c).strip())
            else:
                overlapped.append(c)
            prev_tail = c[-chunk_overlap_chars:]
        return overlapped

    return chunks


@dataclass(frozen=True)
class IngestResult:
    doc_id: str
    doc_version: int
    num_chunks: int
    embedding_dim: int
    content_sha256: str
    changed: bool


_embedder_singleton: Embedder | None = None
_embedder_lock = Lock()


def _get_embedder() -> Embedder:
    global _embedder_singleton
    if _embedder_singleton is not None:
        return _embedder_singleton

    with _embedder_lock:
        if _embedder_singleton is not None:
            return _embedder_singleton

        backend = settings.embeddings_backend
        try:
            if backend == "none":
                _embedder_singleton = NoEmbedder()
            elif backend == "hash":
                _embedder_singleton = HashEmbedder(dim=settings.embedding_dim)
            elif backend == "sentence-transformers":
                _embedder_singleton = SentenceTransformerEmbedder(settings.embeddings_model)
            else:
                _embedder_singleton = HashEmbedder(dim=settings.embedding_dim)
        except Exception as e:  # pragma: no cover
            # Optional dependency might be missing or model download may fail.
            logger.warning("Failed to initialize embedder backend=%s; falling back to hash. error=%s", backend, e)
            _embedder_singleton = HashEmbedder(dim=settings.embedding_dim)

        return _embedder_singleton


def ingest_text(
    *,
    title: str,
    source: str,
    text: str,
    doc_id: str | None = None,
    classification: str | None = None,
    retention: str | None = None,
    tags: str | list[str] | None = None,
    notes: str | None = None,
    schema_fingerprint: str | None = None,
    contract_sha256: str | None = None,
    validation_status: str | None = None,
    validation_errors: list[str] | None = None,
    schema_drifted: bool = False,
) -> IngestResult:
    """Ingest a text blob.

    This method is designed to be replayable/idempotent at the doc_id layer:
    - doc_id defaults to a stable hash of (title, source)
    - ingestion records an ingest_events lineage row capturing content hash and settings
    """

    embedder = _get_embedder()
    doc_id = doc_id or stable_doc_id(title, source)

    # Compute content fingerprint for drift tracking.
    text_bytes = text.encode("utf-8", errors="replace")
    content_sha256 = hashlib.sha256(text_bytes).hexdigest()
    content_bytes = len(text_bytes)

    cls = normalize_classification(classification)
    ret = normalize_retention(retention)
    tag_list = normalize_tags(tags)
    tags_json = json.dumps(tag_list, ensure_ascii=False)

    chunks = chunk_text(text, settings.chunk_size_chars, settings.chunk_overlap_chars)
    chunk_objs: list[Chunk] = []
    for i, c in enumerate(chunks):
        chunk_id = f"{doc_id}__{i:05d}"
        chunk_objs.append(Chunk(chunk_id=chunk_id, doc_id=doc_id, idx=i, text=c))

    embs = embedder.embed([c.text for c in chunk_objs])  # (n, dim)
    if settings.database_url:
        rows = [
            (chunk.chunk_id, int(embs.shape[1]), _vec_to_pgvector_literal(vec))
            for chunk, vec in zip(chunk_objs, embs, strict=True)
        ]
    else:
        rows = [
            (chunk.chunk_id, int(embs.shape[1]), vec.astype(np.float32).tobytes())
            for chunk, vec in zip(chunk_objs, embs, strict=True)
        ]

    with connect(settings.sqlite_path) as conn:
        init_db(conn)

        # Ensure stored embeddings are compatible with the currently configured backend/model.
        # This prevents "silent" retrieval degradation when settings change but the DB is reused.
        # Postgres path currently uses direct schema migrations and skips this SQLite-specific rebuild flow.
        if not settings.database_url:
            ensure_index_compatible(conn, embedder)

        existing = get_doc(conn, doc_id)
        prev_hash = existing.content_sha256 if existing is not None else None
        prev_version = existing.doc_version if existing is not None else 0
        doc_version = prev_version + 1 if existing is not None else 1
        changed = prev_hash is None or prev_hash != content_sha256

        inferred_drifted = bool(schema_drifted)
        if schema_fingerprint:
            prev_events = list_ingest_events(conn, doc_id, limit=1)
            prev_schema = prev_events[0].schema_fingerprint if prev_events else None
            inferred_drifted = inferred_drifted or bool(prev_schema and prev_schema != schema_fingerprint)

        effective_validation_status = validation_status
        if effective_validation_status == "pass" and inferred_drifted:
            effective_validation_status = "warn"
        validation_errors_json = (
            json.dumps(validation_errors or [], ensure_ascii=False) if validation_errors is not None else None
        )

        upsert_doc(
            conn,
            doc_id=doc_id,
            title=title,
            source=source,
            classification=cls,
            retention=ret,
            tags_json=tags_json,
            content_sha256=content_sha256,
            content_bytes=content_bytes,
            num_chunks=len(chunk_objs),
            doc_version=doc_version,
        )

        # Re-ingest replaces doc contents.
        delete_doc_contents(conn, doc_id)
        insert_chunks(conn, chunk_objs)
        insert_embeddings(conn, rows)

        # Lineage artifact for drift/audit.
        evt = IngestEvent(
            event_id=str(uuid.uuid4()),
            doc_id=doc_id,
            doc_version=doc_version,
            ingested_at=int(time.time()),
            content_sha256=content_sha256,
            prev_content_sha256=prev_hash,
            changed=1 if changed else 0,
            num_chunks=len(chunk_objs),
            embedding_backend=settings.embeddings_backend,
            embeddings_model=settings.embeddings_model,
            embedding_dim=int(embs.shape[1]),
            chunk_size_chars=settings.chunk_size_chars,
            chunk_overlap_chars=settings.chunk_overlap_chars,
            schema_fingerprint=schema_fingerprint,
            contract_sha256=contract_sha256,
            validation_status=effective_validation_status,
            validation_errors_json=validation_errors_json,
            schema_drifted=1 if inferred_drifted else 0,
            notes=(notes or None),
        )
        insert_ingest_event(conn, evt)

        conn.commit()

    return IngestResult(
        doc_id=doc_id,
        doc_version=doc_version,
        num_chunks=len(chunk_objs),
        embedding_dim=int(embs.shape[1]),
        content_sha256=content_sha256,
        changed=changed,
    )


def _truncate_cell(value: object, *, max_chars: int) -> str:
    """Normalize a cell value into a bounded string.

    We intentionally keep this simple and deterministic.
    """

    if value is None:
        s = ""
    else:
        s = str(value)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.strip()
    if len(s) > max_chars:
        s = s[: max_chars - 1].rstrip() + "…"
    return s


def _csv_to_text(path: Path) -> tuple[str, str, TabularSnapshot]:
    """Convert a CSV/TSV into a retrieval-friendly text representation.

    Returns: (text, notes)
    """

    import csv

    # Default delimiter based on extension.
    default_delim = "\t" if path.suffix.lower() == ".tsv" else ","

    delimiter = default_delim

    # Read a small sample for dialect sniffing.
    sample: str
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        sample = f.read(4096)
        f.seek(0)

        try:
            sniffed = csv.Sniffer().sniff(sample)
            # Only accept a small set of safe delimiters.
            if sniffed.delimiter in {",", "\t", ";", "|"}:
                delimiter = sniffed.delimiter
        except Exception:
            # Fallback is fine.
            delimiter = default_delim

        reader = csv.reader(f, delimiter=delimiter)
        rows: list[list[str]] = []
        truncated = False
        for i, row in enumerate(reader):
            if i >= _TABULAR_MAX_ROWS:
                truncated = True
                break
            # Cap columns.
            if len(row) > _TABULAR_MAX_COLS:
                row = row[:_TABULAR_MAX_COLS]
                truncated = True
            rows.append([_truncate_cell(c, max_chars=_TABULAR_MAX_CELL_CHARS) for c in row])

    if not rows:
        return "", "csv rows=0", build_snapshot([], [])

    # Determine header.
    header = rows[0]
    data_rows = rows[1:]
    # If header looks empty, synthesize.
    if not any(h.strip() for h in header):
        header = [f"col_{i + 1}" for i in range(len(rows[0]))]
        data_rows = rows

    col_count = len(header)

    # Render as deterministic, paragraph-separated rows.
    # This works well with the chunker (splits on blank lines).
    lines: list[str] = []
    lines.append(f"Spreadsheet ingestion ({path.name})")
    lines.append(f"Format: CSV-like (delimiter='{delimiter}')")
    lines.append(f"Columns ({col_count}): {', '.join(h or f'col_{i + 1}' for i, h in enumerate(header))}")
    lines.append("")

    total_chars = sum(len(x) for x in lines)
    emitted = 0
    emitted_rows: list[dict[str, str]] = []
    for r_idx, row in enumerate(data_rows, start=1):
        if emitted >= _TABULAR_MAX_ROWS:
            truncated = True
            break
        # Normalize row length to header.
        if len(row) < col_count:
            row = row + [""] * (col_count - len(row))
        if len(row) > col_count:
            row = row[:col_count]

        lines.append(f"Row {r_idx}:")
        row_map: dict[str, str] = {}
        for c_idx, (h, v) in enumerate(zip(header, row, strict=False)):
            key = h or f"col_{c_idx + 1}"
            lines.append(f"- {key}: {v}")
            row_map[key] = v
        lines.append("")
        emitted_rows.append(row_map)
        emitted += 1
        total_chars += sum(len(x) for x in lines[-(col_count + 2) :])
        if total_chars >= _TABULAR_MAX_TOTAL_CHARS:
            truncated = True
            break

    if truncated:
        lines.append("(truncated)")

    notes = f"csv rows_emitted={emitted} cols={col_count} truncated={1 if truncated else 0}"
    snapshot = build_snapshot(
        [h or f"col_{i + 1}" for i, h in enumerate(header)],
        emitted_rows,
    )
    return "\n".join(lines).strip() + "\n", notes, snapshot


def _xlsx_to_text(path: Path) -> tuple[str, str, TabularSnapshot]:
    """Convert an XLSX/XLSM workbook into text.

    This uses `openpyxl` (included in the default dependency set for this repo).
    """

    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "XLSX ingestion requires 'openpyxl'. Install it in your environment and retry. "
            "E.g. `uv sync` (default deps) or `uv pip install openpyxl`. "
            f"(import error: {e})"
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    max_sheets = min(len(sheet_names), 5)

    lines: list[str] = []
    lines.append(f"Spreadsheet ingestion ({path.name})")
    lines.append(f"Format: XLSX (sheets={len(sheet_names)}; processed={max_sheets})")
    lines.append("")

    total_chars = sum(len(x) for x in lines)
    truncated = False
    emitted_rows_total = 0

    emitted_rows: list[dict[str, str]] = []
    for s_idx, name in enumerate(sheet_names[:max_sheets], start=1):
        ws = wb[name]

        # Read up to max rows and max cols.
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= _TABULAR_MAX_ROWS:
                truncated = True
                break
            vals = [_truncate_cell(c, max_chars=_TABULAR_MAX_CELL_CHARS) for c in row[:_TABULAR_MAX_COLS]]
            if len(row) > _TABULAR_MAX_COLS:
                truncated = True
            rows.append(vals)
        if not rows:
            continue

        header = rows[0]
        data_rows = rows[1:]
        if not any(h.strip() for h in header):
            header = [f"col_{i + 1}" for i in range(len(rows[0]))]
            data_rows = rows

        col_count = len(header)
        lines.append(f"Sheet {s_idx}: {name}")
        lines.append(f"Columns ({col_count}): {', '.join(h or f'col_{i + 1}' for i, h in enumerate(header))}")
        lines.append("")

        for r_idx, row in enumerate(data_rows, start=1):
            if emitted_rows_total >= _TABULAR_MAX_ROWS:
                truncated = True
                break
            if len(row) < col_count:
                row = row + [""] * (col_count - len(row))
            if len(row) > col_count:
                row = row[:col_count]

            lines.append(f"Row {r_idx}:")
            row_map: dict[str, str] = {}
            for c_idx, (h, v) in enumerate(zip(header, row, strict=False)):
                key = h or f"col_{c_idx + 1}"
                lines.append(f"- {key}: {v}")
                row_map[key] = v
            lines.append("")
            emitted_rows.append(row_map)
            emitted_rows_total += 1

            total_chars += sum(len(x) for x in lines[-(col_count + 2) :])
            if total_chars >= _TABULAR_MAX_TOTAL_CHARS:
                truncated = True
                break

        if total_chars >= _TABULAR_MAX_TOTAL_CHARS:
            truncated = True
            break

    if truncated:
        lines.append("(truncated)")

    notes = (
        f"xlsx sheets={len(sheet_names)} processed={max_sheets} rows_emitted={emitted_rows_total} "
        f"truncated={1 if truncated else 0}"
    )
    # Use headers from the last processed non-empty sheet for fingerprinting.
    # If multiple sheets differ, drift will surface via content/schema changes over time.
    last_headers: list[str] = []
    for line in lines:
        if line.startswith("Columns (") and ":" in line:
            cols = line.split(":", 1)[1].strip()
            if cols:
                last_headers = [c.strip() for c in cols.split(",")]
    snapshot = build_snapshot(last_headers, emitted_rows)
    return "\n".join(lines).strip() + "\n", notes, snapshot


def ingest_file(
    path: str | Path,
    *,
    title: str | None = None,
    source: str | None = None,
    classification: str | None = None,
    retention: str | None = None,
    tags: str | list[str] | None = None,
    notes: str | None = None,
    contract_bytes: bytes | None = None,
) -> IngestResult:
    path = Path(path)
    title = title or path.stem
    # Use filename as the default source so doc_ids are stable across machines.
    source = source or path.name

    suffix = path.suffix.lower()

    if contract_bytes is not None and suffix not in {".csv", ".tsv", ".xlsx", ".xlsm"}:
        raise ValueError("contract_file is only supported for tabular files (.csv/.tsv/.xlsx/.xlsm)")

    if suffix in {".md", ".txt"}:
        text = path.read_text(encoding="utf-8")
        return ingest_text(
            title=title,
            source=source,
            text=text,
            classification=classification,
            retention=retention,
            tags=tags,
            notes=notes,
        )

    if suffix in {".csv", ".tsv"}:
        text, auto_notes, snapshot = _csv_to_text(path)
        if not text.strip():
            raise RuntimeError("No text could be extracted from CSV/TSV")
        contract_sha256: str | None = None
        validation_status: str | None = None
        validation_errors: list[str] | None = None
        if contract_bytes is not None:
            contract, contract_sha256 = load_contract(contract_bytes)
            validation = validate_snapshot(snapshot, contract)
            validation_status = validation.status
            validation_errors = [*validation.errors, *validation.warnings]
            if validation.status == "fail":
                raise ValueError("; ".join(validation.errors) or "Contract validation failed")

        merged_notes = auto_notes if not notes else f"{notes}\n{auto_notes}"
        return ingest_text(
            title=title,
            source=source,
            text=text,
            classification=classification,
            retention=retention,
            tags=tags,
            notes=merged_notes,
            schema_fingerprint=snapshot.schema_fingerprint,
            contract_sha256=contract_sha256,
            validation_status=validation_status,
            validation_errors=validation_errors,
        )

    if suffix in {".xlsx", ".xlsm"}:
        text, auto_notes, snapshot = _xlsx_to_text(path)
        if not text.strip():
            raise RuntimeError("No text could be extracted from XLSX/XLSM")
        xlsx_contract_sha256: str | None = None
        xlsx_validation_status: str | None = None
        xlsx_validation_errors: list[str] | None = None
        if contract_bytes is not None:
            contract, xlsx_contract_sha256 = load_contract(contract_bytes)
            validation = validate_snapshot(snapshot, contract)
            xlsx_validation_status = validation.status
            xlsx_validation_errors = [*validation.errors, *validation.warnings]
            if validation.status == "fail":
                raise ValueError("; ".join(validation.errors) or "Contract validation failed")

        merged_notes = auto_notes if not notes else f"{notes}\n{auto_notes}"
        return ingest_text(
            title=title,
            source=source,
            text=text,
            classification=classification,
            retention=retention,
            tags=tags,
            notes=merged_notes,
            schema_fingerprint=snapshot.schema_fingerprint,
            contract_sha256=xlsx_contract_sha256,
            validation_status=xlsx_validation_status,
            validation_errors=xlsx_validation_errors,
        )

    if suffix == ".pdf":
        # Robust extraction via PyMuPDF; optionally OCR scanned pages with Tesseract.
        res = extract_text_from_pdf(path)
        if not res.text.strip():
            raise RuntimeError("No text could be extracted from PDF (is it scanned? enable OCR_ENABLED=1).")
        auto_notes = f"pdf pages={res.pages} ocr_pages={res.ocr_pages}"
        if res.warnings:
            auto_notes = auto_notes + " warnings=" + ";".join(res.warnings[:5])
        merged_notes = auto_notes if not notes else f"{notes}\n{auto_notes}"
        return ingest_text(
            title=title,
            source=source,
            text=res.text,
            classification=classification,
            retention=retention,
            tags=tags,
            notes=merged_notes,
        )

    raise ValueError(f"Unsupported file type: {path.suffix}. Use .md, .txt, .pdf, .csv, .tsv, .xlsx, .xlsm")
